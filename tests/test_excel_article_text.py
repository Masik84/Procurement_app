from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from app.exports.excel_column_format import excel_value_by_header
from app.utils.excel_export_format import write_openpyxl_dict_sheet
from app.utils.excel_fast_writer import write_excel_table
from app.utils.excel_headers import article_text, is_article_header


class _Column:
    NumberFormatLocal = "General"
    NumberFormat = "General"


class _CellRef:
    def __init__(self, row: int, column: int) -> None:
        self.row = row
        self.column = column


class _Range:
    def __init__(self, worksheet: "_Worksheet", first: _CellRef, last: _CellRef) -> None:
        self.worksheet = worksheet
        self.first = first
        self.last = last
        self._value = None

    @property
    def Value(self):
        return self._value

    @Value.setter
    def Value(self, value) -> None:
        self._value = value
        self.worksheet.writes.append((self.first, self.last, value))


class _Worksheet:
    def __init__(self) -> None:
        self.writes: list[tuple[_CellRef, _CellRef, object]] = []
        self.column_formats: dict[str, _Column] = {}

    def Cells(self, row: int, column: int) -> _CellRef:
        return _CellRef(row, column)

    def Range(self, first: _CellRef, last: _CellRef) -> _Range:
        return _Range(self, first, last)

    def Columns(self, address: str) -> _Column:
        return self.column_formats.setdefault(address, _Column())


class ExcelArticleTextTests(unittest.TestCase):
    def test_article_headers_are_recognized_in_all_export_spellings(self) -> None:
        for header in ("Article", "Supplier Article", "SourceArticle", "supplier_article", "Артикул", "Артикул_2"):
            with self.subTest(header=header):
                self.assertTrue(is_article_header(header))
        self.assertFalse(is_article_header("Product name"))

    def test_article_value_conversion_preserves_leading_zeroes(self) -> None:
        self.assertEqual(article_text("001230"), "001230")
        self.assertEqual(article_text(1230), "1230")
        self.assertEqual(article_text(1230.0), "1230")
        self.assertEqual(article_text("001230.0"), "001230")
        self.assertEqual(excel_value_by_header("Supplier Article", 1230.0), "1230")

    def test_com_table_writer_formats_and_writes_articles_as_text(self) -> None:
        worksheet = _Worksheet()

        write_excel_table(
            worksheet,
            ["SourceArticle", "Product"],
            [["00123", "First"], [456.0, "Second"]],
        )

        self.assertEqual(worksheet.column_formats["A:A"].NumberFormatLocal, "@")
        self.assertEqual(
            worksheet.writes[-1][2],
            (("00123", "First"), ("456", "Second")),
        )

    def test_openpyxl_writer_stores_article_cells_as_text(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        write_openpyxl_dict_sheet(
            worksheet,
            [
                {"SourceArticle": "00123", "Product": "First"},
                {"SourceArticle": 456.0, "Product": "Second"},
                {"SourceArticle": "=FORMULA-LIKE", "Product": "Third"},
            ],
        )

        self.assertEqual(worksheet["A2"].value, "00123")
        self.assertEqual(worksheet["A3"].value, "456")
        self.assertEqual(worksheet["A2"].data_type, "s")
        self.assertEqual(worksheet["A3"].data_type, "s")
        self.assertEqual(worksheet["A2"].number_format, "@")
        self.assertEqual(worksheet["A3"].number_format, "@")
        self.assertEqual(worksheet["A4"].value, "=FORMULA-LIKE")
        self.assertEqual(worksheet["A4"].data_type, "s")

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "article_text.xlsx"
            workbook.save(output_path)
            saved_worksheet = load_workbook(output_path, data_only=False).active
            self.assertEqual(saved_worksheet["A2"].value, "00123")
            self.assertEqual(saved_worksheet["A2"].data_type, "s")
            self.assertEqual(saved_worksheet["A2"].number_format, "@")
            self.assertEqual(saved_worksheet["A4"].value, "=FORMULA-LIKE")
            self.assertEqual(saved_worksheet["A4"].data_type, "s")


if __name__ == "__main__":
    unittest.main()
