from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.utils.parsers import parse_loose_number
from app.utils.text import clean_multi_spaces


class SupplierPriceImporter:
    """Reads supplier price Excel templates into rows for SupplierPriceService."""

    # ВАЖНО: ключи здесь нормализованы через _norm_header():
    # убираем пробелы, запятые, точки, скобки и прочие символы.
    # Так заголовки "Qty, pcs", "Qty pcs", "Qty. pcs", "Volume,L"
    # будут распознаны одинаково.
    HEADER_ALIASES = {
        "materialnumber": "supplier_article",
        "article": "supplier_article",
        "supplierarticle": "supplier_article",
        "material": "product_name",
        "supplierproductname": "product_name",
        "productname": "product_name",
        "pricel": "price",
        "pricelt": "price",
        "pricepack": "price_pack",
        "qtypcs": "qty_pcs",
        "qtypc": "qty_pcs",
        "qtypieces": "qty_pcs",
        "quantitypcs": "qty_pcs",
        "quantity": "qty_pcs",
        "volumel": "volume_l",
        "volumelt": "volume_l",
        "volume": "volume_l",
        "volemul": "volume_l",
    }

    @staticmethod
    def _norm_header(value: Any) -> str:
        text = clean_multi_spaces(str(value or "")).strip().lower()
        return "".join(ch for ch in text if ch.isalnum())

    @staticmethod
    def _clean_text(value: Any) -> str | None:
        if value is None:
            return None
        text = clean_multi_spaces(str(value))
        if not text or text.lower() == "nan":
            return None
        return text

    @staticmethod
    def _num(value: Any):
        if value is None or value == "":
            return None
        return parse_loose_number(value)

    def read_excel(self, file_path: str | Path) -> list[dict]:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []

        columns: dict[int, str] = {}
        for index, header in enumerate(header_row):
            key = self.HEADER_ALIASES.get(self._norm_header(header))
            if key:
                columns[index] = key

        rows: list[dict] = []
        for excel_row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = {
                "import_row_no": excel_row_no,
                "supplier_article": None,
                "product_name": None,
                "price": None,
                "price_pack": None,
                "qty_pcs": None,
                "volume_l": None,
            }
            for index, key in columns.items():
                value = values[index] if index < len(values) else None
                if key in {"supplier_article", "product_name"}:
                    row[key] = self._clean_text(value)
                else:
                    row[key] = self._num(value)

            if any(row.get(k) not in (None, "") for k in ("supplier_article", "product_name", "price", "price_pack", "qty_pcs", "volume_l")):
                rows.append(row)

        wb.close()
        return rows
