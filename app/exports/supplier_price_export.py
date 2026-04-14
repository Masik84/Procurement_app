from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.db.models import CurrentSupplierPrice, PriceHistory, Product, ProductStock, Supplier, SupplierPriceCalculation, TempPriceImport
from app.services.cost_calculation import CostCalculationService


class SupplierPriceExport:
    def __init__(self, session: Session):
        self.session = session
        self.cost_calculation = CostCalculationService(session)

    @staticmethod
    def _to_decimal(value: object) -> Decimal:
        if value is None:
            return Decimal("0")
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))

    @staticmethod
    def _safe_filename(value: str) -> str:
        s = (value or "").strip()
        for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
            s = s.replace(ch, "_")
        return s or "Supplier"

    def _excel_value(self, value: object):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _calc_pack_price(self, price_per_l: object, pack: object):
        if price_per_l is None or pack is None:
            return None
        d_price = self._to_decimal(price_per_l)
        d_pack = self._to_decimal(pack)
        if d_pack == 0:
            return None
        return float((d_price * d_pack).quantize(Decimal("0.0001")))

    def _calc_supplier_full_cost_from_db(self, supplier_id: int, product_id: int, supplier_price: object):
        if supplier_price is None:
            return None

        supplier = self.session.query(Supplier).filter(Supplier.id == supplier_id).first()
        if supplier is None:
            return None

        fx_rate = self.session.query(Supplier.base_currency).filter(Supplier.id == supplier_id).scalar()
        if not fx_rate:
            return None

        from app.services.supplier import SupplierService
        supplier_service = SupplierService(self.session)
        rate_to_rub = supplier_service.get_rate_to_rub(supplier.base_currency)

        if rate_to_rub is None or float(rate_to_rub) == 0:
            return None

        try:
            result = self.cost_calculation.calculate_supplier_costs(
                supplier_id=supplier_id,
                product_id=product_id,
                supplier_price=self._to_decimal(supplier_price),
                fx_rate=self._to_decimal(rate_to_rub),
                currency_code=supplier.base_currency,
            )
            return result.full_cost_msk
        except Exception:
            return None

    def _consider_best_candidate(self, cand_supplier_name, cand_full_cost, cand_date, best1, best2):
        if not cand_supplier_name or cand_full_cost is None:
            return best1, best2

        if best1["price"] is None:
            best1 = {"supplier": cand_supplier_name, "price": cand_full_cost, "date": cand_date}
            return best1, best2

        if self._to_decimal(cand_full_cost) < self._to_decimal(best1["price"]):
            best2 = best1
            best1 = {"supplier": cand_supplier_name, "price": cand_full_cost, "date": cand_date}
            return best1, best2

        if best2["price"] is None:
            best2 = {"supplier": cand_supplier_name, "price": cand_full_cost, "date": cand_date}
            return best1, best2

        if self._to_decimal(cand_full_cost) < self._to_decimal(best2["price"]):
            best2 = {"supplier": cand_supplier_name, "price": cand_full_cost, "date": cand_date}

        return best1, best2

    def _get_best_two_suppliers_for_export(self, current_supplier_id: int, product_id: int, current_supplier_name: str, current_imported_full_cost, current_imported_date):
        best1 = {"supplier": "", "price": None, "date": None}
        best2 = {"supplier": "", "price": None, "date": None}
        seen = set()

        current_supplier = self.session.query(Supplier).filter(Supplier.id == current_supplier_id).first()
        current_supplier_in_rating = bool(current_supplier.rating_calc) if current_supplier else False

        if current_supplier_in_rating:
            best1, best2 = self._consider_best_candidate(current_supplier_name, current_imported_full_cost, current_imported_date, best1, best2)

        seen.add(current_supplier_id)

        current_rows = self.session.query(CurrentSupplierPrice, Supplier).join(
            Supplier, Supplier.id == CurrentSupplierPrice.supplier_id
        ).filter(
            CurrentSupplierPrice.product_id == product_id,
            CurrentSupplierPrice.supplier_id != current_supplier_id,
            CurrentSupplierPrice.price.isnot(None),
            Supplier.rating_calc.is_(True),
        ).all()

        for current_price, supplier in current_rows:
            if supplier.id in seen:
                continue
            v_full_cost = self._calc_supplier_full_cost_from_db(supplier.id, product_id, current_price.price)
            best1, best2 = self._consider_best_candidate(supplier.name, v_full_cost, current_price.last_update, best1, best2)
            seen.add(supplier.id)

        history_subq = self.session.query(
            PriceHistory.supplier_id,
            PriceHistory.product_id,
            PriceHistory.price_date,
            PriceHistory.price,
        ).filter(
            PriceHistory.product_id == product_id,
            PriceHistory.supplier_id != current_supplier_id,
            PriceHistory.price.isnot(None),
        ).order_by(
            PriceHistory.supplier_id.asc(),
            PriceHistory.price_date.desc(),
            PriceHistory.id.desc(),
        ).all()

        latest_history_by_supplier = {}
        for row in history_subq:
            if row.supplier_id not in latest_history_by_supplier:
                latest_history_by_supplier[row.supplier_id] = row

        for supplier_id, row in latest_history_by_supplier.items():
            if supplier_id in seen:
                continue
            supplier = self.session.query(Supplier).filter(Supplier.id == supplier_id, Supplier.rating_calc.is_(True)).first()
            if supplier is None:
                continue
            v_full_cost = self._calc_supplier_full_cost_from_db(supplier.id, product_id, row.price)
            best1, best2 = self._consider_best_candidate(supplier.name, v_full_cost, row.price_date, best1, best2)
            seen.add(supplier.id)

        return best1, best2

    def build_export_dataframe(self, batch_id: str, imported_by: str) -> pd.DataFrame:
        rows = self.session.query(TempPriceImport, SupplierPriceCalculation, Product, ProductStock, Supplier).outerjoin(
            SupplierPriceCalculation,
            (TempPriceImport.batch_id == SupplierPriceCalculation.batch_id)
            & (TempPriceImport.imported_by == SupplierPriceCalculation.imported_by)
            & (TempPriceImport.import_row_no == SupplierPriceCalculation.import_row_no),
        ).outerjoin(
            Product, TempPriceImport.selected_product_id == Product.id
        ).outerjoin(
            ProductStock, TempPriceImport.selected_product_id == ProductStock.product_id
        ).outerjoin(
            Supplier, TempPriceImport.supplier_id == Supplier.id
        ).filter(
            TempPriceImport.batch_id == batch_id,
            TempPriceImport.imported_by == imported_by,
        ).order_by(TempPriceImport.import_row_no.asc()).all()

        out_rows = []

        for temp_row, calc_row, product, stock, supplier in rows:
            current_supplier_name = supplier.name if supplier else ""
            product_id_for_row = temp_row.selected_product_id or 0

            best1 = {"supplier": "", "price": None, "date": None}
            best2 = {"supplier": "", "price": None, "date": None}

            if product_id_for_row > 0:
                best1, best2 = self._get_best_two_suppliers_for_export(
                    current_supplier_id=temp_row.supplier_id,
                    product_id=product_id_for_row,
                    current_supplier_name=current_supplier_name,
                    current_imported_full_cost=calc_row.full_cost_msk if calc_row else None,
                    current_imported_date=calc_row.calc_date if calc_row else None,
                )

            pack_value = product.pack if product is not None else temp_row.new_pack
            price_per_l = temp_row.price
            price_pack_export = temp_row.price_pack if temp_row.price_pack is not None else self._calc_pack_price(price_per_l, pack_value)

            transit_total = None
            if stock is not None:
                transit_total = self._to_decimal(stock.transit_qty) + self._to_decimal(stock.is_confirmed_order_qty)

            out_rows.append({
                "Supplier Article": temp_row.supplier_article or "",
                "Supplier Product Name": temp_row.product_name or "",
                "Our Product Name": product.name if product else "",
                "Pack": self._excel_value(pack_value),
                "Price, L": self._excel_value(price_per_l),
                "Price (Pack)": self._excel_value(price_pack_export),
                "Currency": calc_row.currency_code if calc_row else (supplier.base_currency if supplier else ""),
                "Cost Novo withVAT": self._excel_value(calc_row.cost_novo_wvat if calc_row else None),
                "Full Cost Msk": self._excel_value(calc_row.full_cost_msk if calc_row else None),
                "Дистр цена": self._excel_value(stock.distr_price if stock else None),
                "Промо цена": self._excel_value(stock.promo_price if stock else None),
                "curr LPC": self._excel_value(stock.lpc if stock else None),
                "curr Landed cost": self._excel_value(stock.landed_cost if stock else None),
                "Best Suppl": best1["supplier"],
                "Best full Price, L": self._excel_value(best1["price"]),
                "last update Best1": best1["date"],
                "Best Suppl 2": best2["supplier"],
                "Best full Price, L 2": self._excel_value(best2["price"]),
                "last update Best2": best2["date"],
                "Stock": self._excel_value(stock.stock_qty if stock else None),
                "Transit": self._excel_value(transit_total),
                "Purchase Order": self._excel_value(stock.order_qty if stock else None),
                "Order IS": self._excel_value(stock.is_order_qty if stock else None),
                "Stock IS": self._excel_value(stock.is_stock_qty if stock else None),
                "Reserve cust": self._excel_value(stock.reserve_qty if stock else None),
                "Damaged": self._excel_value(stock.markdown_qty if stock else None),
            })

        return pd.DataFrame(out_rows)

    def export_calculated(self, batch_id: str, imported_by: str, supplier_id: int, output_path: str | Path | None = None) -> Path:
        supplier = self.session.query(Supplier).filter(Supplier.id == supplier_id).first()
        supplier_name = self._safe_filename(supplier.name if supplier else "Supplier")

        if output_path is None:
            filename = f"CostCalc_{supplier_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            output_path = Path(filename)
        else:
            output_path = Path(output_path)

        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        df = self.build_export_dataframe(batch_id=batch_id, imported_by=imported_by)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.book["Sheet1"]

            header_fill_gray = PatternFill("solid", fgColor="CDCDCD")
            header_fill_red = PatternFill("solid", fgColor="C00000")
            header_fill_blue = PatternFill("solid", fgColor="00B0F0")
            header_fill_green = PatternFill("solid", fgColor="92D050")
            header_fill_dark_blue = PatternFill("solid", fgColor="215C98")
            white_font = Font(color="FFFFFF", bold=True)
            bold_font = Font(bold=True)
            center = Alignment(horizontal="center", vertical="top", wrap_text=True)

            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = center

            for col in range(1, 10):
                ws.cell(1, col).fill = header_fill_gray

            for col in range(10, 14):
                ws.cell(1, col).fill = header_fill_red
                ws.cell(1, col).font = white_font

            for col in range(14, 17):
                ws.cell(1, col).fill = header_fill_blue

            for col in range(17, 20):
                ws.cell(1, col).fill = header_fill_green

            for col in range(20, 23):
                ws.cell(1, col).fill = header_fill_dark_blue
                ws.cell(1, col).font = white_font

            for col in range(23, 25):
                ws.cell(1, col).fill = header_fill_red
                ws.cell(1, col).font = white_font

            for col in range(25, 27):
                ws.cell(1, col).fill = header_fill_dark_blue
                ws.cell(1, col).font = white_font

            for col in ["A"]:
                ws.column_dimensions[col].width = 16
            for col in ["B", "C"]:
                ws.column_dimensions[col].width = 31.14
            ws.column_dimensions["D"].width = 10
            for col in ["N", "Q"]:
                ws.column_dimensions[col].width = 16.14
            ws.column_dimensions["P"].width = 9.43
            ws.column_dimensions["S"].width = 9.86
            for col in ["T", "U", "V", "W", "X", "Y", "Z"]:
                ws.column_dimensions[col].width = 8.14

            for col in ["E", "F"]:
                for cell in ws[col][1:]:
                    cell.number_format = "#,##0.00"

            for col in ["H", "I", "J", "K", "L", "M", "O", "R"]:
                for cell in ws[col][1:]:
                    cell.number_format = '#,##0 "$"'

            for col in ["P", "S"]:
                for cell in ws[col][1:]:
                    cell.number_format = "dd/mm/yy;@"

            for col in ["T", "U", "V", "W", "X", "Y", "Z"]:
                for cell in ws[col][1:]:
                    cell.number_format = '#,##0;[Red]-#,##0;"-"'

            ws.auto_filter.ref = "A1:Z1"
            ws.freeze_panes = "L2"

        return output_path