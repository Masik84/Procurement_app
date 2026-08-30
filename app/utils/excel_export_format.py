from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Mapping, Sequence

from app.utils.excel_headers import article_text, display_header, is_article_header
from app.utils.excel_fast_writer import write_excel_table
from app.utils.excel_format_rules import FORMATS, set_number_format_safe

XL_CENTER = -4108
XL_LEFT = -4131
XL_RIGHT = -4152
XL_VCENTER = -4160

FONT_NAME = "Aptos Narrow"
FONT_SIZE = 11
HEADER_ROW_HEIGHT = 45

TEXT_FORMAT = FORMATS.TEXT
DATE_FORMAT_LOCAL = FORMATS.DATE
INTEGER_FORMAT_LOCAL = FORMATS.INTEGER
DECIMAL_FORMAT_LOCAL = FORMATS.DECIMAL_2
DECIMAL4_FORMAT_LOCAL = FORMATS.DECIMAL_4
MONEY_FORMAT_LOCAL = FORMATS.MONEY_RUB
GENERAL_FORMAT = FORMATS.GENERAL

DEFAULT_HEADER_COLOR = (205, 205, 205)
SUPPLIER_HEADER_COLOR = (146, 208, 80)
COST_HEADER_COLOR = (0, 176, 240)
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)


def rgb(r: int, g: int, b: int) -> int:
    return r + g * 256 + b * 65536


def excel_column_letter(col_num: int) -> str:
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def header_map(headers: Sequence[str]) -> dict[str, int]:
    return {str(header): idx + 1 for idx, header in enumerate(headers)}


def col_letter(header_to_index: Mapping[str, int], header: str) -> str | None:
    idx = header_to_index.get(header)
    return excel_column_letter(idx) if idx else None


def apply_base_table_style(ws, headers_count: int) -> None:
    ws.Cells.Font.Name = FONT_NAME
    ws.Cells.Font.Size = FONT_SIZE
    last_col = excel_column_letter(headers_count)
    header_range = ws.Range(f"A1:{last_col}1")
    header_range.Font.Name = FONT_NAME
    header_range.Font.Size = FONT_SIZE
    header_range.Font.Bold = True
    header_range.WrapText = True
    header_range.HorizontalAlignment = XL_CENTER
    header_range.VerticalAlignment = XL_VCENTER
    ws.Rows(1).RowHeight = HEADER_ROW_HEIGHT


def normalize_header(header: str) -> str:
    text = " ".join(str(header or "").strip().lower().replace("\n", " ").split())
    # Dynamic exports often add suffixes: Cost Novo_1, Full Cost Msk_2, etc.
    if "_" in text:
        base, suffix = text.rsplit("_", 1)
        if suffix.isdigit():
            text = base
    return text


def is_date_header(header: str) -> bool:
    h = normalize_header(header)
    return h in {"дата", "price date", "date"} or h.endswith(" date")


def is_text_header(header: str) -> bool:
    if is_article_header(header):
        return True
    h = normalize_header(header)
    return h in {
        "id", "менеджер", "клиент", "customer product name", "our product name",
        "supplier", "supplier article", "supplier product name", "currency", "comments", "has customs", "via novo",
        "excise duty", "final supplier", "product name", "manager name", "customer name",
        "article", "brand", "family", "категория abc",
    }


INTEGER_HEADERS = {
    "qty, pcs",
    "volume, l",
    "pack",
    "stockqty",
    "transitqty",
    "markdownqty",
    "reserveqty",
    "reserveecommqty",
    "orderqty",
    "confirmedqty",
    "remainsqty",
    "lpc",
    # Existing exported quantity-like headers kept as integer format.
    "qty",
    "quantity",
    "importrowno",
    "stock",
    "transit",
    "purchase order",
    "order is",
    "stock is",
    "reserve cust",
    "reserve e-comm",
    "damaged",
    "к быстрому заказу, шт",
    "к заказу, шт",
}


def is_integer_header(header: str) -> bool:
    return normalize_header(header) in INTEGER_HEADERS


def is_money_header(header: str) -> bool:
    h = normalize_header(header)
    return h in {"final price", "price rub", "cost novo withvat", "cost novo with vat", "full cost msk"}


def is_decimal4_header(header: str) -> bool:
    h = normalize_header(header)
    return h in {
        "supplier price, l", "supplier price", "price, l", "price, pack", "cost per l", "price per l",
        "cost novo", "cost novo wvat", "fx rate",
    }


def is_decimal_header(header: str) -> bool:
    h = normalize_header(header)
    return (
        is_decimal4_header(header)
        or is_money_header(header)
        or h in {
            "fx markup %", "fx markup abs", "transport", "re-export", "agent fee", "bank fee", "customs fee",
            "additional customs", "storage", "move novo", "move msk", "marking",
            "ср.продажи мес", "safe stock (st), mnth", "safe stock (st+tr), mnth", "safe stock (+ord), mnth",
            "дистр цена", "промо цена", "к быстрому заказу, л", "к заказу, л",
        }
    )


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text or text in {"-", "—"}:
        return None
    text = text.replace("\u00a0", " ").replace("₽", "").replace("руб.", "").replace("руб", "")
    text = text.replace(" ", "")
    if "," in text and "." in text:
        # 1,234.56 -> 1234.56, 1.234,56 -> 1234.56
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return value


def excel_cell_value(header: str, value: Any) -> Any:
    if value is None or value == "":
        return ""
    if is_article_header(header):
        return article_text(value)
    if is_date_header(header):
        return parse_date(value)
    if is_decimal_header(header) or is_integer_header(header):
        number = parse_decimal(value)
        if number is None or number == 0:
            return ""
        if is_integer_header(header):
            return int(number.quantize(Decimal("1")))
        return float(number)
    if isinstance(value, Decimal):
        return float(value) if value != 0 else ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    return value


def column_format_local(header: str) -> str:
    if is_date_header(header):
        return DATE_FORMAT_LOCAL
    if is_text_header(header):
        return TEXT_FORMAT
    if is_integer_header(header):
        return INTEGER_FORMAT_LOCAL
    if is_money_header(header):
        return MONEY_FORMAT_LOCAL
    if is_decimal4_header(header):
        return DECIMAL4_FORMAT_LOCAL
    if is_decimal_header(header):
        return DECIMAL_FORMAT_LOCAL
    return GENERAL_FORMAT


def apply_column_formats(ws, headers: Sequence[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        letter = excel_column_letter(idx)
        fmt = column_format_local(display_header(header))
        if fmt == TEXT_FORMAT:
            set_number_format_safe(ws.Columns(f"{letter}:{letter}"), TEXT_FORMAT, TEXT_FORMAT)
        elif fmt == GENERAL_FORMAT:
            set_number_format_safe(ws.Columns(f"{letter}:{letter}"), GENERAL_FORMAT, GENERAL_FORMAT)
        else:
            set_number_format_safe(ws.Columns(f"{letter}:{letter}"), GENERAL_FORMAT, fmt)


def write_table(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    write_excel_table(
        ws,
        headers,
        rows,
        header_getter=display_header,
        value_getter=lambda row, header, col_index: excel_cell_value(
            str(header),
            row[col_index] if col_index < len(row) else "",
        ),
    )


def set_widths_by_headers(ws, headers: Sequence[str], widths: Mapping[str, float], default_width: float = 10.0) -> None:
    hm = header_map(headers)
    for header in headers:
        letter = col_letter(hm, header)
        if letter:
            ws.Columns(f"{letter}:{letter}").ColumnWidth = widths.get(header, default_width)


def color_headers(ws, headers: Sequence[str], color_map: Mapping[str, tuple[int, int, int]], default_color: tuple[int, int, int] = DEFAULT_HEADER_COLOR) -> None:
    hm = header_map(headers)
    last_col = excel_column_letter(len(headers))
    ws.Range(f"A1:{last_col}1").Interior.Color = rgb(*default_color)
    ws.Range(f"A1:{last_col}1").Font.Color = rgb(*BLACK)
    for header, color in color_map.items():
        letter = col_letter(hm, header)
        if letter:
            ws.Range(f"{letter}1").Interior.Color = rgb(*color)


# ---------- High-level Excel export helpers used by all project exporters ----------

def write_dict_table(ws, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    """Write headers and dict rows with column-name based value conversion."""
    write_excel_table(
        ws,
        headers,
        rows,
        header_getter=display_header,
        value_getter=lambda row, header, _col_index: excel_cell_value(str(header), row.get(header, "")),
    )


def apply_standard_table_format(
    ws,
    headers: Sequence[str],
    *,
    widths: Mapping[str, float] | None = None,
    color_map: Mapping[str, tuple[int, int, int]] | None = None,
    apply_filter: bool = True,
    default_width: float = 12.0,
) -> None:
    """Apply one shared style/format policy to COM Excel worksheets."""
    if not headers:
        return
    apply_base_table_style(ws, len(headers))
    color_headers(ws, headers, color_map or {})
    apply_column_formats(ws, headers)
    if widths:
        set_widths_by_headers(ws, headers, widths, default_width=default_width)
    else:
        for idx in range(1, len(headers) + 1):
            letter = excel_column_letter(idx)
            ws.Columns(f"{letter}:{letter}").ColumnWidth = default_width
    if apply_filter:
        try:
            last_col = excel_column_letter(len(headers))
            ws.Range(f"A1:{last_col}1").AutoFilter(1)
        except Exception:
            pass


def write_and_format_table(
    ws,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]] | Sequence[Mapping[str, Any]],
    *,
    widths: Mapping[str, float] | None = None,
    color_map: Mapping[str, tuple[int, int, int]] | None = None,
    apply_filter: bool = True,
) -> None:
    """Write a complete table and apply shared Excel formatting."""
    if rows and isinstance(rows[0], Mapping):
        write_dict_table(ws, headers, rows)  # type: ignore[arg-type]
    else:
        write_table(ws, headers, rows)  # type: ignore[arg-type]
    apply_standard_table_format(ws, headers, widths=widths, color_map=color_map, apply_filter=apply_filter)


def openpyxl_cell_value(header: str, value: Any) -> Any:
    """Same conversion policy for openpyxl-based exports."""
    return excel_cell_value(header, value)


def openpyxl_number_format(header: str) -> str:
    if is_date_header(header):
        return FORMATS.DATE
    if is_text_header(header):
        return FORMATS.TEXT
    if is_integer_header(header):
        return FORMATS.INTEGER
    if is_money_header(header):
        return FORMATS.MONEY_RUB
    if is_decimal4_header(header):
        return FORMATS.DECIMAL_4
    if is_decimal_header(header):
        return FORMATS.DECIMAL_2
    return FORMATS.GENERAL


def write_openpyxl_dict_sheet(ws, rows: Sequence[Mapping[str, Any]], *, widths: Mapping[str, float] | None = None) -> None:
    """Write an openpyxl sheet using the same project-wide column policy."""
    if not rows:
        return
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = list(rows[0].keys())
    ws.append([display_header(h) for h in headers])
    for row in rows:
        ws.append([openpyxl_cell_value(h, row.get(h, "")) for h in headers])

    header_fill = PatternFill("solid", fgColor="CDCDCD")
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    for col_index, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_index)
        fmt = openpyxl_number_format(header)
        for cell in ws[col_letter]:
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=(cell.row == 1))
            if cell.row > 1:
                cell.number_format = fmt
                if is_article_header(header):
                    # openpyxl otherwise treats strings beginning with "=" as
                    # formulas even when the display number format is text.
                    cell.value = article_text(cell.value)
                    cell.data_type = "s"
        if widths and header in widths:
            ws.column_dimensions[col_letter].width = widths[header]
        else:
            max_len = max(len(str(c.value or "")) for c in ws[col_letter])
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
