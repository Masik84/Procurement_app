from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import uuid
from pathlib import Path
from typing import Any
from math import isnan

from openpyxl import Workbook

from PySide6.QtCore import QFile, Qt, QUrl, QTimer, QEvent, QPoint
from PySide6.QtGui import QDesktopServices, QFont
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QMenu,
    QMessageBox,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from PySide6.QtUiTools import QUiLoader

from app.db.db import SessionLocal
from app.db.models import (
    Product,
    TempIsImport,
    TempStockImport,
    TempSupplierOrdersImport,
)
from app.exports.product_stock_exporter import ProductStockExporter
from app.services.product_matching_service import ProductMatchingService
from app.services.product_stock_service import ProductStockService
from app.utils.batch import get_current_username
from app.utils.parsers import parse_loose_number
from app.utils.text import clean_multi_spaces
from app.ui.table_style import *
from app.utils.output_headers import display_headers

BASE_DIR = Path(__file__).resolve().parents[2]
UI_PATH = BASE_DIR / "app" / "ui" / "windows" / "stock_supplier_orders.ui"


@dataclass(slots=True)
class ColumnDef:
    key: str
    header: str
    editable: bool = False
    kind: str = "text"  # text | product_combo | brand_combo | checkbox | indicator


MODE_STOCK = "stock"
MODE_ORDERS = "orders"
MODE_IS = "is"


COLUMN_DEFS: dict[str, list[ColumnDef]] = {
    MODE_STOCK: [
        ColumnDef("selected_product_id", "Our product", kind="product_combo"),
        ColumnDef("source_article", "Source article", editable=True),
        ColumnDef("source_sku", "SKU", editable=True),
        ColumnDef("source_product_name", "Source product", editable=True),
        ColumnDef("abc_category", "Категория ABC", editable=True),
        ColumnDef("stock_qty", "Stock qty", editable=True),
        ColumnDef("transit_qty", "Transit qty", editable=True),
        ColumnDef("markdown_qty", "Markdown qty", editable=True),
        ColumnDef("reserve_qty", "Reserve qty", editable=True),
        ColumnDef("reserve_ecomm_qty", "Reserve E-Comm", editable=True),
        ColumnDef("has_lpc_warning", "LPC err", kind="indicator"),
        ColumnDef("lpc", "LPC", editable=True),
        ColumnDef("landed_cost", "Landed cost", editable=True),
        ColumnDef("distr_price", "Distr price", editable=True),
        ColumnDef("promo_price", "Promo price", editable=True),
        ColumnDef("new_product_name", "Product name (new)", editable=True),
        ColumnDef("new_brand", "Brand (new)", kind="brand_combo"),
        ColumnDef("new_pack", "Pack (new)", editable=True),
        ColumnDef("new_is_excise", "Excise (new)", kind="checkbox"),
    ],
    MODE_ORDERS: [
        ColumnDef("selected_product_id", "Our product", kind="product_combo"),
        ColumnDef("source_article", "Source article", editable=True),
        ColumnDef("source_product_name", "Source product", editable=True),
        ColumnDef("abc_category", "Категория ABC", editable=True),
        ColumnDef("order_qty", "Order qty", editable=True),
        ColumnDef("new_product_name", "Product name (new)", editable=True),
        ColumnDef("new_brand", "Brand (new)", kind="brand_combo"),
        ColumnDef("new_pack", "Pack (new)", editable=True),
        ColumnDef("new_is_excise", "Excise (new)", kind="checkbox"),
    ],
    MODE_IS: [
        ColumnDef("selected_product_id", "Our product", kind="product_combo"),
        ColumnDef("source_article", "Source article", editable=True),
        ColumnDef("source_product_name", "Source product", editable=True),
        ColumnDef("confirmed_qty", "Confirmed qty", editable=True),
        ColumnDef("remains_qty", "Remains qty", editable=True),
        ColumnDef("stock_qty", "Stock qty", editable=True),
        ColumnDef("new_product_name", "Product name (new)", editable=True),
        ColumnDef("new_brand", "Brand (new)", kind="brand_combo"),
        ColumnDef("new_pack", "Pack (new)", editable=True),
        ColumnDef("new_is_excise", "Excise (new)", kind="checkbox"),
    ],
}


NUMERIC_FIELDS = {
    "stock_qty", "markdown_qty", "reserve_qty", "reserve_ecomm_qty", "lpc", "landed_cost", "distr_price", "promo_price",
    "transit_qty", "order_qty", "confirmed_qty", "remains_qty", "new_pack",
}


def load_ui(ui_path: Path):
    loader = QUiLoader()
    ui_file = QFile(str(ui_path))
    if not ui_file.open(QFile.ReadOnly):
        raise RuntimeError(f"Не удалось открыть UI: {ui_path}")
    try:
        widget = loader.load(ui_file)
    finally:
        ui_file.close()
    if widget is None:
        raise RuntimeError(f"Не удалось загрузить UI: {ui_path}")
    return widget


class ProductStockPage(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = load_ui(UI_PATH)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.ui)

        self.table = self.ui.table
        self._updating_table = False
        self._batch_id = ""
        self._imported_by = get_current_username()
        self._current_file_path = ""
        self._mode = MODE_STOCK
        self._product_name_cache: dict[int, str] = {}
        self._editing_table_cell = False

        self.setup_ui()
        self.setup_connections()
        self.start_new_batch()
        self.cleanup_old_temp_rows()
        self.refresh_filters()
        self.apply_mode(MODE_STOCK)
        self.clear_message()

    def get_session(self):
        return SessionLocal()

    def cleanup_old_temp_rows(self):
        with self.get_session() as session:
            service = ProductStockService(session)
            service.cleanup_old_temp_rows(imported_by=self._imported_by)
            session.commit()

    def eventFilter(self, watched, event):
        if isinstance(watched, QComboBox):
            row_id = watched.property("row_id")
            role = watched.property("combo_role")
            if row_id and role and event.type() in {QEvent.FocusIn, QEvent.MouseButtonPress}:
                if role == "brand_combo":
                    self.populate_brand_combo(watched, keep_current=True)

        return super().eventFilter(watched, event)

    def setup_ui(self):
        setup_data_table(self.table, sorting=True)
        self.table.horizontalHeader().setSectionsMovable(False)
        from app.utils.gui_table_actions import install_standard_table_context_menu
        install_standard_table_context_menu(self, self.table)

        # font = QFont("Tahoma", 10)
        for widget in (self.ui.line_RowsLoaded, self.ui.line_TotalQty, self.ui.line_RowsError):
            widget.setReadOnly(True)
            widget.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            # widget.setFont(font)

        self.clear_table()

    def setup_connections(self):
        self.ui.radio_Stock.toggled.connect(lambda checked: checked and self.apply_mode(MODE_STOCK))
        self.ui.radio_Orders.toggled.connect(lambda checked: checked and self.apply_mode(MODE_ORDERS))
        self.ui.radio_IS.toggled.connect(lambda checked: checked and self.apply_mode(MODE_IS))

        self.ui.line_FindProduct.textChanged.connect(self.refresh_current_product_combo)
        self.ui.cbo_FindBrand.currentTextChanged.connect(self.refresh_current_product_combo)

        self.ui.btn_Import.clicked.connect(self.import_file)
        self.ui.btn_Save.clicked.connect(self.save_all)
        self.ui.btn_Reset.clicked.connect(self.reset_all)
        self.table.cellDoubleClicked.connect(self.start_cell_edit)
        self.table.itemChanged.connect(self.on_item_changed)


    def show_message(self, text: str):
        self.ui.label_msg.setText(text)
        self.ui.label_msg.setProperty("active", True)
        self.ui.label_msg.style().unpolish(self.ui.label_msg)
        self.ui.label_msg.style().polish(self.ui.label_msg)
        self.ui.label_msg.setVisible(True)
        QTimer.singleShot(10000, self.clear_message)

    def clear_message(self):
        self.ui.label_msg.setText("")
        self.ui.label_msg.setProperty("active", False)
        self.ui.label_msg.style().unpolish(self.ui.label_msg)
        self.ui.label_msg.style().polish(self.ui.label_msg)
        self.ui.label_msg.setVisible(True)

    def show_error_message(self, text: str):
        self.clear_message()

        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Ошибка")
        msg.setText(text)

        copy_btn = msg.addButton("Copy", QMessageBox.ActionRole)
        msg.addButton(QMessageBox.Ok)

        msg.exec()

        if msg.clickedButton() == copy_btn:
            QApplication.clipboard().setText(text)

    def start_new_batch(self):
        self._batch_id = f"PS_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{uuid.uuid4().hex[:6]}"
        self._imported_by = get_current_username()
        self._current_file_path = ""

    def apply_mode(self, mode: str):
        self._mode = mode
        self.clear_table()
        self.refresh_counters()

    def clear_table(self):
        self._updating_table = True
        try:
            self.table.blockSignals(True)
            self.table.clear()
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
        finally:
            self.table.blockSignals(False)
            self._updating_table = False

    def refresh_filters(self):
        with self.get_session() as session:
            brands = (
                session.query(Product.brand)
                .filter(Product.brand.isnot(None), Product.brand != "")
                .distinct()
                .order_by(Product.brand.asc())
                .all()
            )
        self.ui.cbo_FindBrand.blockSignals(True)
        self.ui.cbo_FindBrand.clear()
        self.ui.cbo_FindBrand.addItem("-")
        for row in brands:
            self.ui.cbo_FindBrand.addItem(row[0])
        self.ui.cbo_FindBrand.blockSignals(False)

    def _temp_model(self):
        return {
            MODE_STOCK: TempStockImport,
            MODE_ORDERS: TempSupplierOrdersImport,
            MODE_IS: TempIsImport,
        }[self._mode]

    def _product_match_mode(self) -> str:
        return MODE_IS if self._mode == MODE_IS else MODE_STOCK

    def _get_filtered_products(self):
        with self.get_session() as session:
            query = session.query(Product).filter(Product.name.isnot(None), Product.name != "")
            brand = self.ui.cbo_FindBrand.currentText().strip()
            find_text = self.ui.line_FindProduct.text().strip().lower()
            if brand and brand != "-":
                query = query.filter(Product.brand == brand)
            products = query.order_by(Product.name.asc()).all()
            if find_text:
                products = [p for p in products if find_text in (p.name or "").lower()]
            return products

    def _get_brand_values(self):
        with self.get_session() as session:
            rows = (
                session.query(Product.brand)
                .filter(Product.brand.isnot(None), Product.brand != "")
                .distinct()
                .order_by(Product.brand.asc())
                .all()
            )
        return [r[0] for r in rows]

    def _query_mode_rows(self, session):
        model = self._temp_model()
        return (
            session.query(model)
            .filter(model.batch_id == self._batch_id, model.imported_by == self._imported_by)
            .order_by(model.import_row_no.asc(), model.id.asc())
            .all()
        )

    def load_table(self):
        with self.get_session() as session:
            rows = self._query_mode_rows(session)
            data = [self._row_to_dict(row) for row in rows]
            product_ids = sorted({int(row.selected_product_id) for row in rows if row.selected_product_id is not None})
            self._product_name_cache = {}
            if product_ids:
                products = session.query(Product.id, Product.name).filter(Product.id.in_(product_ids)).all()
                self._product_name_cache = {int(pid): (name or "") for pid, name in products}
        if not data:
            self.clear_table()
        else:
            self.display_table(data)
        self.refresh_counters()

    def _row_to_dict(self, row) -> dict[str, Any]:
        base = {
            "id": row.id,
            "selected_product_id": row.selected_product_id,
            "source_article": row.source_article,
            "source_product_name": row.source_product_name,
            "new_product_name": row.new_product_name,
            "new_brand": row.new_brand,
            "new_pack": row.new_pack,
            "new_is_excise": bool(row.new_is_excise) if row.new_is_excise is not None else False,
        }
        if isinstance(row, TempStockImport):
            base.update({
                "source_sku": row.source_sku,
                "abc_category": row.abc_category,
                "stock_qty": row.stock_qty,
                "transit_qty": row.transit_qty,
                "markdown_qty": row.markdown_qty,
                "reserve_qty": row.reserve_qty,
                "reserve_ecomm_qty": getattr(row, "reserve_ecomm_qty", 0),
                "has_lpc_warning": bool(row.has_lpc_warning),
                "lpc": row.lpc,
                "landed_cost": row.landed_cost,
                "distr_price": row.distr_price,
                "promo_price": row.promo_price,
            })
        elif isinstance(row, TempSupplierOrdersImport):
            base.update({
                "abc_category": row.abc_category,
                "order_qty": row.order_qty,
            })
        elif isinstance(row, TempIsImport):
            base.update({
                "confirmed_qty": row.confirmed_qty,
                "remains_qty": row.remains_qty,
                "stock_qty": row.stock_qty,
            })
        return base

    def _safe_float(self, value: Any) -> float | None:
        if value is None:
            return None
        try:
            value = float(value)
        except (TypeError, ValueError):
            return None
        if isnan(value):
            return None
        return value

    def _format_int_like(self, value: Any, blank_zero: bool = False) -> str:
        num = self._safe_float(value)
        if num is None:
            return ""
        rounded = int(round(num))
        if blank_zero and rounded == 0:
            return ""
        return f"{rounded:,}".replace(",", " ")

    def _format_decimal1(self, value: Any, blank_zero: bool = False) -> str:
        num = self._safe_float(value)
        if num is None:
            return ""
        if blank_zero and abs(num) < 1e-12:
            return ""
        return f"{num:,.1f}".replace(",", " ")

    def _format_cell_value(self, field_name: str, value: Any) -> str:
        if field_name in {"stock_qty", "transit_qty", "markdown_qty", "reserve_qty", "reserve_ecomm_qty", "order_qty", "confirmed_qty", "remains_qty"}:
            return self._format_int_like(value, blank_zero=True)
        if field_name in {"lpc", "landed_cost", "distr_price", "promo_price"}:
            return self._format_decimal1(value, blank_zero=True)
        if field_name == "new_pack":
            num = self._safe_float(value)
            if num is None:
                return ""
            if abs(num - round(num)) < 1e-12:
                return self._format_int_like(num)
            return self._format_decimal1(num)
        if value is None:
            return ""
        text = str(value).strip()
        return "" if text.lower() == "nan" else text

    def _apply_table_column_layout(self):
        for col_index, col in enumerate(COLUMN_DEFS[self._mode]):
            if col.key == "selected_product_id":
                self.table.setColumnWidth(col_index, 150)
            elif col.key == "source_article":
                self.table.setColumnWidth(col_index, 130)
            elif col.key == "source_sku":
                self.table.setColumnWidth(col_index, 95)
            elif col.key == "source_product_name":
                self.table.setColumnWidth(col_index, 110)
            elif col.key == "abc_category":
                self.table.setColumnWidth(col_index, 110)

    def display_table(self, data: list[dict[str, Any]]):
        self._updating_table = True
        try:
            columns = COLUMN_DEFS[self._mode]
            self.table.blockSignals(True)
            self.table.clear()
            self.table.setColumnCount(len(columns))
            self.table.setHorizontalHeaderLabels([c.header for c in columns])
            self.table.setRowCount(len(data))

            font = QFont("Tahoma", 10)
            self.table.setFont(font)
            self.table.horizontalHeader().setFont(font)

            for row_index, row_data in enumerate(data):
                row_id = row_data["id"]
                for col_index, col in enumerate(columns):
                    value = row_data.get(col.key)
                    if col.kind == "checkbox":
                        self.table.setCellWidget(row_index, col_index, self._build_checkbox(row_id, col.key, bool(value)))
                    elif col.kind == "indicator":
                        self.table.setCellWidget(row_index, col_index, self._build_indicator(bool(value)))
                    elif col.kind == "product_combo":
                        text = self._product_name_cache.get(int(value), "") if value is not None else ""
                        self.table.setItem(row_index, col_index, self._build_display_item(row_id, col.key, text))
                    elif col.kind == "brand_combo":
                        self.table.setItem(row_index, col_index, self._build_display_item(row_id, col.key, self._format_cell_value(col.key, value)))
                    else:
                        item = QTableWidgetItem(self._format_cell_value(col.key, value))
                        item.setData(Qt.UserRole, row_id)
                        flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
                        if col.editable:
                            flags |= Qt.ItemIsEditable
                        item.setFlags(flags)
                        if col.key in NUMERIC_FIELDS:
                            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        else:
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        self.table.setItem(row_index, col_index, item)

            self.table.resizeColumnsToContents()
            self._apply_table_column_layout()
        finally:
            self.table.blockSignals(False)
            self._updating_table = False

    def start_cell_edit(self, row: int, column: int):
        if self._updating_table:
            return

        columns = COLUMN_DEFS[self._mode]
        if row < 0 or row >= self.table.rowCount() or column < 0 or column >= len(columns):
            return

        col = columns[column]
        if col.kind not in {"product_combo", "brand_combo"}:
            return

        row_id = None
        for check_col in range(self.table.columnCount()):
            item = self.table.item(row, check_col)
            if item is not None:
                row_id = item.data(Qt.UserRole)
                if row_id is not None:
                    break

        if row_id is None:
            return

        if col.kind == "product_combo":
            combo = self._build_product_combo(row_id, self._get_selected_product_id(row_id))
            combo.activated.connect(
                lambda _, r=row, rid=row_id, c=combo, key=col.key: self.finish_product_edit(r, rid, c, key)
            )

            self.table.setCellWidget(row, column, combo)
            combo.setFocus()
            QTimer.singleShot(0, combo.showPopup)

        elif col.kind == "brand_combo":
            combo = self._build_brand_combo(row_id, self._get_brand_text(row_id), self._get_brand_values())
            combo.activated.connect(
                lambda _, r=row, rid=row_id, c=combo, key=col.key: self.finish_brand_edit(r, rid, c, key)
            )

            if combo.lineEdit() is not None:
                combo.lineEdit().returnPressed.connect(
                    lambda r=row, rid=row_id, c=combo, key=col.key: self.finish_brand_edit(r, rid, c, key)
                )

            self.table.setCellWidget(row, column, combo)
            combo.setFocus()
            if combo.lineEdit() is not None:
                combo.lineEdit().selectAll()

    def _build_display_item(self, row_id: int, field_name: str, value: str):
        item = QTableWidgetItem(value)
        item.setData(Qt.UserRole, row_id)
        item.setData(Qt.UserRole + 1, field_name)
        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        if field_name in NUMERIC_FIELDS:
            item.setTextAlignment(Qt.AlignCenter)
        else:
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        return item

    def _get_product_name_by_id(self, product_id):
        if not product_id:
            return ""
        try:
            product_id = int(product_id)
        except (TypeError, ValueError):
            return ""
        cached = self._product_name_cache.get(product_id)
        if cached is not None:
            return cached
        with self.get_session() as session:
            product = session.query(Product).filter(Product.id == product_id).first()
            name = product.name or "" if product else ""
            self._product_name_cache[product_id] = name
            return name

    def _get_selected_product_id(self, row_id: int):
        with self.get_session() as session:
            row = self._get_row_by_id(session, row_id)
            return row.selected_product_id if row else None

    def _get_brand_text(self, row_id: int):
        with self.get_session() as session:
            row = self._get_row_by_id(session, row_id)
            return row.new_brand or "" if row else ""

    def finish_product_edit(self, row: int, row_id: int, combo: QComboBox, field_name: str):
        value = combo.currentData()
        try:
            value = int(value)
        except (TypeError, ValueError):
            value = None

        current_col = self.table.currentColumn()
        current_row = self.table.currentRow()
        self.update_temp_field(row_id, field_name, value, reload=False)
        self._updating_table = True
        self.table.removeCellWidget(row, self._column_index_by_key(field_name))
        self.table.setItem(
            row,
            self._column_index_by_key(field_name),
            self._build_display_item(row_id, field_name, combo.currentText().strip()),
        )
        self._updating_table = False
        if value is not None:
            self._product_name_cache[int(value)] = combo.currentText().strip()
        self.table.setCurrentCell(current_row if current_row >= 0 else row, current_col if current_col >= 0 else self._column_index_by_key(field_name))

    def finish_brand_edit(self, row: int, row_id: int, combo: QComboBox, field_name: str):
        if self._editing_table_cell:
            return

        self._editing_table_cell = True
        try:
            text = clean_multi_spaces(combo.currentText())
            text = text.upper() if text else None

            current_col = self.table.currentColumn()
            current_row = self.table.currentRow()

            self.update_temp_field(
                row_id,
                field_name,
                text,
                reload=False,
                clear_selected=True,
            )

            self._updating_table = True
            self.table.removeCellWidget(row, self._column_index_by_key(field_name))
            self.table.setItem(
                row,
                self._column_index_by_key(field_name),
                self._build_display_item(row_id, field_name, text or ""),
            )
            self._updating_table = False

            self.table.setCurrentCell(
                current_row if current_row >= 0 else row,
                current_col if current_col >= 0 else self._column_index_by_key(field_name),
            )
        finally:
            self._editing_table_cell = False

    def _column_index_by_key(self, field_name: str) -> int:
        for index, col in enumerate(COLUMN_DEFS[self._mode]):
            if col.key == field_name:
                return index
        return -1

    def _commit_open_editors(self):
        for row in range(self.table.rowCount()):
            for column in range(self.table.columnCount()):
                widget = self.table.cellWidget(row, column)
                if not isinstance(widget, QComboBox):
                    continue

                field_name = COLUMN_DEFS[self._mode][column].key

                row_id = None
                for check_col in range(self.table.columnCount()):
                    item = self.table.item(row, check_col)
                    if item is not None:
                        row_id = item.data(Qt.UserRole)
                        if row_id is not None:
                            break

                if row_id is None:
                    continue

                if field_name == "new_brand":
                    self.finish_brand_edit(row, row_id, widget, field_name)
                elif field_name == "selected_product_id":
                    self.finish_product_edit(row, row_id, widget, field_name)
                
    def _build_indicator(self, checked: bool):
        checkbox = QCheckBox()
        checkbox.setChecked(checked)
        checkbox.setEnabled(False)
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(checkbox)
        layout.setAlignment(Qt.AlignCenter)
        return container

    def _build_checkbox(self, row_id: int, field_name: str, checked: bool):
        checkbox = QCheckBox()
        checkbox.setChecked(checked)
        checkbox.toggled.connect(lambda state, rid=row_id, fn=field_name: self.update_temp_field(rid, fn, bool(state), reload=False, clear_selected=True))
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(checkbox)
        layout.setAlignment(Qt.AlignCenter)
        return container

    def _build_brand_combo(self, row_id: int, value: str | None, brands: list[str]):
        combo = QComboBox()
        combo.setEditable(True)
        combo.setInsertPolicy(QComboBox.NoInsert)
        combo.setProperty("row_id", row_id)
        combo.setProperty("combo_role", "brand_combo")
        # Важно: для Brand (new) не ставим eventFilter.
        # Иначе при клике/фокусе combo заново заполняется через populate_brand_combo(),
        # из-за чего выбранный бренд может не успеть сохраниться и визуально "исчезает".
        # Поведение должно быть как в supplier_prices_page: combo создается один раз,
        # выбранное значение фиксируется в finish_brand_edit().
        self.populate_brand_combo(combo, keep_current=False, current_text=value or "")
        return combo

    def populate_brand_combo(self, combo: QComboBox, keep_current: bool, current_text: str = ""):
        brand_value = combo.currentText().strip() if keep_current else current_text
        brands = self._get_brand_values()

        combo.blockSignals(True)
        combo.clear()
        combo.addItem("")
        if brands:
            combo.addItems(brands)
        if brand_value and combo.findText(brand_value) < 0:
            combo.addItem(brand_value)
        combo.setCurrentText(brand_value)
        combo.blockSignals(False)

    def _build_product_combo(self, row_id: int, selected_product_id: int | None):
        combo = QComboBox()
        combo.setProperty("row_id", row_id)
        combo.setProperty("combo_role", "product_combo")
        combo.setToolTip("Выберите продукт из базы")
        self.populate_product_combo(
            combo,
            row_id=row_id,
            keep_current=False,
            selected_product_id=selected_product_id,
        )
        return combo

    def populate_product_combo(
        self,
        combo: QComboBox,
        row_id: int,
        keep_current: bool,
        selected_product_id: int | None = None,
    ):
        current_id = combo.currentData() if keep_current else selected_product_id

        if current_id is not None:
            try:
                current_id = int(current_id)
            except (TypeError, ValueError):
                current_id = None

        products = self._get_filtered_products()

        combo.blockSignals(True)
        combo.clear()
        combo.addItem("", None)

        for product in products:
            combo.addItem(product.name, int(product.id))

        if current_id is not None and combo.findData(current_id) < 0:
            with self.get_session() as session:
                product = session.query(Product).filter(Product.id == current_id).first()
                if product is not None:
                    combo.addItem(product.name, int(product.id))

        idx = combo.findData(current_id)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        else:
            combo.setCurrentIndex(0)

        combo.blockSignals(False)

    def on_item_changed(self, item: QTableWidgetItem):
        if self._updating_table:
            return
        row_id = item.data(Qt.UserRole)
        if row_id is None:
            return
        col = COLUMN_DEFS[self._mode][item.column()]
        if not col.editable:
            return
        value: Any = item.text().strip()
        if col.key in NUMERIC_FIELDS:
            value = parse_loose_number(value)
        elif value == "":
            value = None

        clear_selected = col.key in {"new_product_name", "new_pack"}
        self.update_temp_field(row_id, col.key, value, reload=False, clear_selected=clear_selected)

        if col.key in {"source_article", "source_product_name"}:
            self.try_auto_match_row(row_id)
        elif col.key in {"new_product_name", "new_brand", "new_pack", "new_is_excise"}:
            self.refresh_counters()

    def _normalize_new_product_text(self, field_name: str, value: Any) -> Any:
        if field_name not in {"new_product_name", "new_brand"}:
            return value
        if value is None:
            return None
        text = clean_multi_spaces(str(value))
        return text.upper() if text else None

    def _get_row_by_id(self, session, row_id: int):
        return session.query(self._temp_model()).filter(self._temp_model().id == row_id).first()

    def update_temp_field(self, row_id: int, field_name: str, value: Any, reload: bool = False, clear_selected: bool = False):
        if self._updating_table:
            return
        value = self._normalize_new_product_text(field_name, value)
        try:
            with self.get_session() as session:
                row = self._get_row_by_id(session, row_id)
                if row is None:
                    return
                setattr(row, field_name, value)
                if clear_selected and value not in (None, ""):
                    row.selected_product_id = None
                if field_name in {"new_product_name", "new_brand", "new_pack"} and value not in (None, ""):
                    if row.new_is_excise is None:
                        row.new_is_excise = False
                if field_name == "selected_product_id" and value is not None:
                    row.new_product_name = None
                    row.new_brand = None
                    row.new_pack = None
                    row.new_is_excise = None
                session.commit()
        except Exception as e:
            self.show_error_message(str(e))
            return

        if reload:
            self.load_table()
        else:
            self.refresh_counters()

    def try_auto_match_row(self, row_id: int):
        try:
            with self.get_session() as session:
                row = self._get_row_by_id(session, row_id)
                if row is None or row.selected_product_id is not None:
                    return
                matcher = ProductMatchingService(session)
                if self._product_match_mode() == MODE_IS:
                    product = matcher.find_is_product(row.source_article, row.source_product_name)
                else:
                    product = matcher.find_stock_product(row.source_article, row.source_product_name)
                if product is not None:
                    row.selected_product_id = product.id
                    session.commit()
                    self.load_table()
                else:
                    self.refresh_counters()
        except Exception as e:
            self.show_error_message(str(e))

    def refresh_current_product_combo(self):
        row = self.table.currentRow()
        if row < 0:
            return

        product_col = self._column_index_by_key("selected_product_id")
        combo = self.table.cellWidget(row, product_col)
        if isinstance(combo, QComboBox):
            row_id = combo.property("row_id")
            if row_id is not None:
                self.populate_product_combo(
                    combo,
                    row_id=int(row_id),
                    keep_current=True,
                )

    def refresh_counters(self):
        with self.get_session() as session:
            model = self._temp_model()
            rows = self._query_mode_rows(session)
            row_count = len(rows)
            error_count = sum(1 for row in rows if row.selected_product_id is None)
            total_qty = 0
            for row in rows:
                if model is TempStockImport:
                    total_qty += float(row.stock_qty or 0) + float(row.transit_qty or 0) + float(row.markdown_qty or 0) + float(row.reserve_qty or 0) + float(getattr(row, "reserve_ecomm_qty", 0) or 0)
                elif model is TempSupplierOrdersImport:
                    total_qty += float(row.order_qty or 0)
                else:
                    total_qty += float(row.confirmed_qty or 0) + float(row.remains_qty or 0) + float(row.stock_qty or 0)

        self.ui.line_RowsLoaded.setText(self._format_int_like(row_count))
        self.ui.line_RowsError.setText(self._format_int_like(error_count))
        self.ui.line_TotalQty.setText(self._format_int_like(total_qty))

    def import_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel файл", "", "Excel files (*.xls *.xlsx *.xlsm)")
        if not file_path:
            return
        try:
            self._current_file_path = file_path
            with self.get_session() as session:
                runner = ProductStockService(session)
                self.start_new_batch()
                if self._mode == MODE_STOCK:
                    runner.import_stock(file_path=file_path, imported_by=self._imported_by, batch_id=self._batch_id)
                elif self._mode == MODE_ORDERS:
                    runner.import_supplier_orders(file_path=file_path, imported_by=self._imported_by, batch_id=self._batch_id)
                else:
                    runner.import_is(file_path=file_path, imported_by=self._imported_by, batch_id=self._batch_id)
                session.commit()

            self.load_table()
            self.offer_save_issue_file()
            self.show_message("Импорт завершен")
        except Exception as e:
            self.show_error_message(str(e))

    def _default_issue_filename(self) -> str:
        suffix = {
            MODE_STOCK: "Stock_Errors",
            MODE_ORDERS: "SupplierOrders_Errors",
            MODE_IS: "IS_Errors",
        }[self._mode]
        return f"{suffix}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    def _collect_issue_sheets(self) -> list[tuple[str, list[dict[str, Any]]]]:
        with self.get_session() as session:
            sheets: list[tuple[str, list[dict[str, Any]]]] = []
            if self._mode == MODE_STOCK:
                rows = (
                    session.query(TempStockImport)
                    .filter(
                        TempStockImport.batch_id == self._batch_id,
                        TempStockImport.imported_by == self._imported_by,
                        TempStockImport.selected_product_id.is_(None),
                    )
                    .order_by(TempStockImport.import_row_no.asc(), TempStockImport.id.asc())
                    .all()
                )
                product_errors = [{
                    "ImportRowNo": row.import_row_no,
                    "SourceArticle": row.source_article,
                    "SourceSKU": row.source_sku,
                    "SourceProductName": row.source_product_name,
                    "Comment": "Не найден SelectedProductID. Требуется сопоставление или создание нового продукта.",
                } for row in rows]
                if product_errors:
                    sheets.append(("Product errors", product_errors))

                warn_rows = (
                    session.query(TempStockImport)
                    .filter(
                        TempStockImport.batch_id == self._batch_id,
                        TempStockImport.imported_by == self._imported_by,
                        TempStockImport.has_lpc_warning.is_(True),
                    )
                    .order_by(TempStockImport.import_row_no.asc(), TempStockImport.id.asc())
                    .all()
                )
                lpc_warnings = [{
                    "ImportRowNo": row.import_row_no,
                    "SourceArticle": row.source_article,
                    "SourceSKU": row.source_sku,
                    "SourceProductName": row.source_product_name,
                    "StockQty": self._format_int_like(row.stock_qty, blank_zero=True),
                    "TransitQty": self._format_int_like(getattr(row, "transit_qty", None), blank_zero=True),
                    "MarkdownQty": self._format_int_like(row.markdown_qty, blank_zero=True),
                    "ReserveQty": self._format_int_like(row.reserve_qty, blank_zero=True),
                    "ReserveECommQty": self._format_int_like(getattr(row, "reserve_ecomm_qty", 0), blank_zero=True),
                    "LPC": self._format_decimal1(row.lpc, blank_zero=True),
                    "Comment": "Есть остаток, но LPC в файле пустой или 0. Итоговый LPC рассчитывается по БД продаж.",
                } for row in warn_rows]
                if lpc_warnings:
                    sheets.append(("LPC warnings", lpc_warnings))

            elif self._mode == MODE_ORDERS:
                rows = (
                    session.query(TempSupplierOrdersImport)
                    .filter(
                        TempSupplierOrdersImport.batch_id == self._batch_id,
                        TempSupplierOrdersImport.imported_by == self._imported_by,
                        TempSupplierOrdersImport.selected_product_id.is_(None),
                    )
                    .order_by(TempSupplierOrdersImport.import_row_no.asc(), TempSupplierOrdersImport.id.asc())
                    .all()
                )
                data = [{
                    "ImportRowNo": row.import_row_no,
                    "SourceArticle": row.source_article,
                    "SourceProductName": row.source_product_name,
                    "OrderQty": self._format_int_like(row.order_qty, blank_zero=True),
                    "Comment": "Не найден SelectedProductID. Требуется сопоставление или создание нового продукта.",
                } for row in rows]
                if data:
                    sheets.append(("Product errors", data))

            else:
                rows = (
                    session.query(TempIsImport)
                    .filter(
                        TempIsImport.batch_id == self._batch_id,
                        TempIsImport.imported_by == self._imported_by,
                        TempIsImport.selected_product_id.is_(None),
                    )
                    .order_by(TempIsImport.import_row_no.asc(), TempIsImport.id.asc())
                    .all()
                )
                data = [{
                    "ImportRowNo": row.import_row_no,
                    "SourceArticle": row.source_article,
                    "SourceProductName": row.source_product_name,
                    "ConfirmedQty": self._format_int_like(row.confirmed_qty, blank_zero=True),
                    "RemainsQty": self._format_int_like(row.remains_qty, blank_zero=True),
                    "StockQty": self._format_int_like(row.stock_qty, blank_zero=True),
                    "Comment": "Не найден SelectedProductID. Требуется сопоставление или создание нового продукта.",
                } for row in rows]
                if data:
                    sheets.append(("Product errors", data))
            return sheets

    def _save_issue_workbook(self, output_path: str, sheets: list[tuple[str, list[dict[str, Any]]]]):
        wb = Workbook()
        first = True
        for title, rows in sheets:
            ws = wb.active if first else wb.create_sheet(title=title[:31])
            ws.title = title[:31]
            first = False

            headers = list(rows[0].keys()) if rows else []
            if headers:
                ws.append(headers)
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])

            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)

            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

        wb.save(output_path)

    def _open_issue_save_dialog(self, sheets: list[tuple[str, list[dict[str, Any]]]]):
        dialog = QFileDialog(self, "Сохранить файл с ошибками", str(BASE_DIR / self._default_issue_filename()), "Excel files (*.xlsx)")
        dialog.setAcceptMode(QFileDialog.AcceptSave)
        dialog.setFileMode(QFileDialog.AnyFile)
        dialog.setOption(QFileDialog.DontUseNativeDialog, False)

        if dialog.exec() != QFileDialog.Accepted:
            return

        selected_files = dialog.selectedFiles()
        if not selected_files:
            return

        output_path = selected_files[0]
        if not output_path:
            return

        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        self._save_issue_workbook(output_path, sheets)

    def offer_save_issue_file(self):
        sheets = self._collect_issue_sheets()
        if not sheets:
            return

        answer = QMessageBox.question(
            self,
            "Ошибки импорта",
            "Сохранить файл с ошибками?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if answer != QMessageBox.Yes:
            return

        QTimer.singleShot(200, lambda s=sheets: self._open_issue_save_dialog(s))

    def save_all(self):
        self._commit_open_editors()

        try:
            with self.get_session() as session:
                from app.utils.gui_table_actions import apply_pending_table_deletes_to_db
                apply_pending_table_deletes_to_db(session, self)
                rows = self._query_mode_rows(session)
                if not rows:
                    self.show_error_message("Нет данных для сохранения. Сначала импортируйте файл или добавьте строки.")
                    return
                matched_count = sum(1 for row in rows if row.selected_product_id is not None)
                if matched_count == 0:
                    self.show_error_message("В текущем импорте нет ни одной строки с SelectedProductID.")
                    return

                runner = ProductStockService(session)
                if self._mode == MODE_STOCK:
                    stats = runner.save_stock(self._batch_id, self._imported_by)
                    msg = "Данные по остаткам успешно сохранены."
                elif self._mode == MODE_ORDERS:
                    stats = runner.save_supplier_orders(self._batch_id, self._imported_by)
                    msg = "Данные по заказам поставщиков успешно сохранены."
                else:
                    stats = runner.save_is(self._batch_id, self._imported_by)
                    msg = "Данные IS успешно сохранены."
                session.commit()

            self.start_new_batch()
            self.clear_table()
            self.refresh_counters()
            self.show_message(msg)
        except Exception as e:
            self.show_error_message(str(e))

    def reset_all(self):
        if QMessageBox.question(self, "Подтверждение", "Сбросить все данные текущего импорта?") != QMessageBox.Yes:
            return
        try:
            with self.get_session() as session:
                service = ProductStockService(session)
                if self._mode == MODE_STOCK:
                    service.delete_stock_rows(self._batch_id, self._imported_by)
                elif self._mode == MODE_ORDERS:
                    service.delete_supplier_orders_rows(self._batch_id, self._imported_by)
                else:
                    service.delete_is_rows(self._batch_id, self._imported_by)
                session.commit()
            self.ui.line_FindProduct.clear()
            self.ui.cbo_FindBrand.setCurrentIndex(0)
            self.start_new_batch()
            self.clear_table()
            self.refresh_counters()
            self.show_message("Форма очищена")
        except Exception as e:
            self.show_error_message(str(e))

